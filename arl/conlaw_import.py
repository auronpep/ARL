from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from arl.io import write_jsonl, write_yaml


REQUIRED_HEADERS = [
    "Question Number",
    "Question",
    "Answer Choice A",
    "Answer Choice B",
    "Answer Choice C",
    "Answer Choice D",
    "Correct Answer",
    "Answer Explanation",
    "Wrong Answer Explanation",
]


STRATEGY_SIGNALS = [
    {
        "mechanic": "federal_police_power_bait",
        "cards": ["CONLAW-ACTOR-FIRST-01", "CONLAW-FEDERAL-POLICE-POWER-BAIT-01"],
        "patterns": ["federal police power", "no federal police power", "congress doesn't have police"],
        "shape": "actor_source_power",
        "axis": "actor/source-of-power",
    },
    {
        "mechanic": "spending_power",
        "cards": ["CONLAW-SPENDING-POWER-01"],
        "patterns": ["federal funds", "tax and spend", "spending power", "general welfare"],
        "shape": "actor_source_power",
        "axis": "spending hook vs bait power",
    },
    {
        "mechanic": "classification_equal_protection",
        "cards": ["CONLAW-CLASSIFICATION-FIRST-01"],
        "patterns": ["classification", "equal protection", "treats", "discriminates", "classifies"],
        "shape": "classification_equal_protection",
        "axis": "classification plus actor",
    },
    {
        "mechanic": "federal_equal_protection_analogue",
        "cards": ["CONLAW-FEDERAL-EP-ANALOGUE-01"],
        "patterns": ["federal statute", "federal law", "congress", "fifth amendment due process"],
        "shape": "classification_equal_protection",
        "axis": "classification plus actor",
        "requires_any": ["equal protection", "classification", "classifies", "discriminates"],
    },
    {
        "mechanic": "pi14_bait",
        "cards": ["CONLAW-PI14-BAIT-01"],
        "patterns": ["fourteenth amendment privileges", "14th amendment privileges", "privileges or immunities clause of the fourteenth"],
        "shape": "clause_home",
        "axis": "clause strength",
    },
    {
        "mechanic": "state_action_threshold",
        "cards": ["CONLAW-STATE-ACTION-FIRST-01"],
        "patterns": ["state action", "private", "public function", "nexus"],
        "shape": "threshold",
        "axis": "threshold before merits",
    },
    {
        "mechanic": "property_interest_trigger",
        "cards": ["CONLAW-PROPERTY-INTEREST-TRIGGER-01"],
        "patterns": ["property interest", "public employment", "tenure", "for cause"],
        "shape": "threshold",
        "axis": "entitlement source",
    },
    {
        "mechanic": "speech_permit_discretion",
        "cards": ["CONLAW-SPEECH-PERMIT-DISCRETION-01"],
        "patterns": ["permit", "license", "demonstrate", "parade", "official discretion"],
        "shape": "threshold",
        "axis": "official discretion",
    },
    {
        "mechanic": "same_sex_marriage_ep",
        "cards": ["CONLAW-SAME-SEX-MARRIAGE-EP-01"],
        "patterns": ["same-sex marriage", "same sex marriage"],
        "shape": "classification_equal_protection",
        "axis": "same-sex marriage disadvantage",
    },
    {
        "mechanic": "contracts_clause_predate",
        "cards": ["CONLAW-CONTRACTS-CLAUSE-PREDATE-01"],
        "patterns": ["contracts clause", "obligations of contracts", "predates", "preexisting contract"],
        "shape": "clause_home",
        "axis": "contract date before statute date",
    },
    {
        "mechanic": "federal_enumerated_power",
        "cards": ["CONLAW-FEDERAL-ENUMERATED-POWER-01"],
        "patterns": ["commerce clause", "necessary and proper", "enumerated power", "congress can act"],
        "shape": "actor_source_power",
        "axis": "enumerated federal power",
    },
]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def infer_strategy_fields(text: str, choices: dict[str, str] | None = None, answer: str | None = None) -> dict[str, Any]:
    haystack = text.lower()
    mechanics: list[str] = []
    mechanic_names: list[str] = []
    shapes: list[str] = []
    axes: list[str] = []

    for signal in STRATEGY_SIGNALS:
        if any(req not in haystack for req in signal.get("requires_all", [])):
            continue
        if signal.get("requires_any") and not any(req in haystack for req in signal["requires_any"]):
            continue
        if any(pattern in haystack for pattern in signal["patterns"]):
            mechanic_names.append(signal["mechanic"])
            mechanics.extend(signal["cards"])
            shapes.append(signal["shape"])
            axes.append(signal["axis"])

    dominant_choice = ""
    if choices:
        for letter, choice_text in choices.items():
            if answer and letter == answer:
                continue
            choice_lower = choice_text.lower()
            if "federal police power" in choice_lower:
                dominant_choice = letter
                if "federal_police_power_bait" not in mechanic_names:
                    mechanic_names.append("federal_police_power_bait")
                    mechanics.extend(["CONLAW-ACTOR-FIRST-01", "CONLAW-FEDERAL-POLICE-POWER-BAIT-01"])
                    shapes.append("actor_source_power")
                    axes.append("actor/source-of-power")
                break

    seen = set()
    expected_mechanic_ids = [m for m in mechanics if not (m in seen or seen.add(m))]
    return {
        "dominant_trap_choice": dominant_choice,
        "dominant_trap_mechanic": mechanic_names[0] if mechanic_names else "",
        "expected_mechanic_ids": expected_mechanic_ids,
        "question_shape": shapes[0] if shapes else "unclassified",
        "expected_axis": axes[0] if axes else "",
    }


def row_to_question(row: dict[str, Any]) -> dict[str, Any]:
    question_number = int(row["Question Number"])
    choices = {
        "A": _clean(row["Answer Choice A"]),
        "B": _clean(row["Answer Choice B"]),
        "C": _clean(row["Answer Choice C"]),
        "D": _clean(row["Answer Choice D"]),
    }
    answer = _clean(row["Correct Answer"]).upper()
    text = " ".join(
        [
            _clean(row["Question"]),
            " ".join(choices.values()),
            _clean(row["Answer Explanation"]),
            _clean(row["Wrong Answer Explanation"]),
        ]
    )
    inferred = infer_strategy_fields(text, choices=choices, answer=answer)

    question = {
        "id": f"CONLAW-SET1-{question_number:03d}",
        "subject": "CONSTITUTIONAL_LAW",
        "topic": "",
        "call": "",
        "stem": _clean(row["Question"]),
        "choices": choices,
        "answer": answer,
        "dominant_trap_choice": inferred["dominant_trap_choice"],
        "dominant_trap_mechanic": inferred["dominant_trap_mechanic"],
        "dominant_trap_pick_rate": 0,
        "expected_phase": "",
        "expected_axis": inferred["expected_axis"],
        "expected_dispositive_fact": "",
        "expected_mechanic_ids": inferred["expected_mechanic_ids"],
        "question_shape": inferred["question_shape"],
        "surface_cluster": "",
        "transfer_cluster": inferred["dominant_trap_mechanic"],
        "difficulty_band": "dev",
        "source": {
            "file": "ConLaw_Set1.xlsx",
            "sheet": "Constitutional Law",
            "question_number": question_number,
        },
        "private_notes": {
            "answer_explanation": _clean(row["Answer Explanation"]),
            "wrong_answer_explanation": _clean(row["Wrong Answer Explanation"]),
        },
    }
    return question


def read_xlsx_rows(path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]
    headers = [_clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"missing workbook headers: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(_clean(value) for value in row.values()):
            rows.append(row)
    return rows


def workbook_to_questions(path: str | Path) -> list[dict[str, Any]]:
    return [row_to_question(row) for row in read_xlsx_rows(path, "Constitutional Law")]


def write_question_outputs(questions: list[dict[str, Any]], out_dir: str | Path) -> dict[str, int]:
    out = Path(out_dir)
    all_rows = list(questions)
    holdout_size = min(30, max(1, round(len(all_rows) * 0.2))) if all_rows else 0
    dev_rows = all_rows[:-holdout_size] if holdout_size else all_rows
    holdout_rows = all_rows[-holdout_size:] if holdout_size else []
    for question in holdout_rows:
        question["difficulty_band"] = "holdout"

    answer_key = [{"id": q["id"], "answer": q["answer"]} for q in all_rows]
    choice_forensics = []
    for q in all_rows:
        for letter, text in q["choices"].items():
            choice_forensics.append(
                {
                    "question_id": q["id"],
                    "choice": letter,
                    "text": text,
                    "is_correct": letter == q["answer"],
                    "dominant_trap_choice": letter == q.get("dominant_trap_choice"),
                    "dominant_trap_mechanic": q.get("dominant_trap_mechanic", ""),
                    "raw_wrong_answer_explanation": q["private_notes"]["wrong_answer_explanation"] if letter != q["answer"] else "",
                    "raw_answer_explanation": q["private_notes"]["answer_explanation"] if letter == q["answer"] else "",
                }
            )

    write_jsonl(out / "questions_set1.jsonl", all_rows)
    write_jsonl(out / "questions_dev.jsonl", dev_rows)
    write_jsonl(out / "questions_holdout.jsonl", holdout_rows)
    write_jsonl(out / "answer_key_set1.jsonl", answer_key)
    write_jsonl(out / "choice_forensics_set1.jsonl", choice_forensics)
    return {
        "questions": len(all_rows),
        "dev": len(dev_rows),
        "holdout": len(holdout_rows),
        "answer_key": len(answer_key),
        "choice_forensics": len(choice_forensics),
    }


def extract_strategy_candidates(md_path: str | Path) -> dict[str, Any]:
    text = Path(md_path).read_text(encoding="utf-8")
    items = []
    for idx, signal in enumerate(STRATEGY_SIGNALS, start=1):
        matched = []
        for pattern in signal["patterns"]:
            match = re.search(re.escape(pattern), text, flags=re.IGNORECASE)
            if match:
                line_no = text[: match.start()].count("\n") + 1
                matched.append({"pattern": pattern, "line": line_no})
        if not matched:
            continue
        items.append(
            {
                "inbox_id": f"INBOX-CONLAW-SET1-{idx:04d}",
                "source_type": "strategy_markdown",
                "source_ref": "Con_Law_Study_Tactics.md",
                "matched_signals": matched,
                "candidate_trigger": signal["patterns"][0],
                "candidate_move": signal["axis"],
                "candidate_trap": signal["mechanic"],
                "candidate_student_script": "",
                "proposed_card_type": "routing_or_trap_mechanic",
                "merge_candidate": signal["cards"][0] if signal["cards"] else "",
                "promote_to_pack": False,
                "notes": "Private candidate extracted from strategy markdown; needs real-question evidence before pack mutation.",
            }
        )
    return {"version": "0.1", "source": "Con_Law_Study_Tactics.md", "items": items}


def import_conlaw_set(xlsx_path: str | Path, md_path: str | Path, out_dir: str | Path) -> dict[str, Any]:
    questions = workbook_to_questions(xlsx_path)
    counts = write_question_outputs(questions, out_dir)
    candidates = extract_strategy_candidates(md_path)
    write_yaml(Path(out_dir) / "strategy_mechanics_candidates.yaml", candidates)

    answer_counts: dict[str, int] = {}
    shape_counts: dict[str, int] = {}
    mechanic_tagged = 0
    dominant_trap_tagged = 0
    for question in questions:
        answer_counts[question["answer"]] = answer_counts.get(question["answer"], 0) + 1
        shape_counts[question["question_shape"]] = shape_counts.get(question["question_shape"], 0) + 1
        mechanic_tagged += int(bool(question["expected_mechanic_ids"]))
        dominant_trap_tagged += int(bool(question["dominant_trap_mechanic"]))

    return {
        **counts,
        "answer_counts": dict(sorted(answer_counts.items())),
        "shape_counts": dict(sorted(shape_counts.items())),
        "questions_with_expected_mechanics": mechanic_tagged,
        "questions_with_dominant_trap_mechanic": dominant_trap_tagged,
        "strategy_candidates": len(candidates["items"]),
    }

